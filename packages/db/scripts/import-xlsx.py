#!/usr/bin/env python3
"""
Import Hotel KinTop's real catalog from the two source workbooks into a single
committed JSON file the Prisma seed loads verbatim.

    python packages/db/scripts/import-xlsx.py

Inputs  (committed for provenance):
    packages/db/data/source/Hotel Kintop Wine Item Stock.xlsx
    packages/db/data/source/Hotel KinTop Kitchen Item List (Filled).xlsx
Output  (committed; the seed reads this — never hand-edit):
    packages/db/data/seed-data.json

Data model
==========
Wine/spirit sheet — one row per product:
  * Spirit (row has any 25..750 ml pour price):
      1 Ingredient  baseUnit=ml, barcode on the *bottle*, openingStock = "MM" (ml)
      1 MenuItem per populated pour price  ("<name> <size>ml"), each a Recipe of
      <size> ml of the bottle. Prices are non-linear, so every pour is its own
      priced unit (50 ml = 1 shot; 2 shots = the 100 ml tier).
  * Discrete (row priced only in "Bottle/Can" — beer, soft drinks, cigarettes):
      1 Ingredient  baseUnit=pcs, openingStock = "QTY" (count)
      1 MenuItem    barcode on the *item*, Recipe = 1 pcs.
  All bar items: category=bar, station=bar, menuGroup = sheet category,
  priced on dine_in_bar + takeaway.

Kitchen sheet — one row per dish:
      1 MenuItem    category=food, station=kitchen, menuGroup = sheet category,
      no barcode (the scanner is bar-only); each populated ingredient cell → a
      Recipe row. Priced on dine_in_restaurant + takeaway + room_service.
  Ingredient columns become Ingredients (unit inferred from the header:
  Egg/Sausages/Cheese = pcs; the rest = g), openingStock 0.

Costs and reorder levels are 0 everywhere (the sheets carry no purchase cost;
weighted-average cost is established later via goods-received).
"""

import json
import sys
from pathlib import Path

import openpyxl

# --- paths -----------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = SCRIPT_DIR.parent / "data"
SOURCE_DIR = DATA_DIR / "source"
WINE_XLSX = SOURCE_DIR / "Hotel Kintop Wine Item Stock.xlsx"
KITCHEN_XLSX = SOURCE_DIR / "Hotel KinTop Kitchen Item List (Filled).xlsx"
OUT_JSON = DATA_DIR / "seed-data.json"

POUR_SIZES = [750, 400, 200, 100, 50, 25]  # ml columns, largest first

BAR_CHANNELS = ["dine_in_bar", "takeaway"]
KITCHEN_CHANNELS = ["dine_in_restaurant", "takeaway", "room_service"]

# Kitchen ingredient columns → (clean ingredient name, base unit). Keyed by the
# stripped header text exactly as it appears in the sheet ("Nooldes" is the
# sheet's own spelling of Noodles; "koththu" is the chopped-roti base).
KITCHEN_ING = {
    "koththu": ("Kottu Roti", "g"),
    "Nooldes": ("Noodles", "g"),
    "Rice (g)": ("Rice", "g"),
    "Egg": ("Egg", "pcs"),
    "Sausages": ("Sausage", "pcs"),
    "Chicken": ("Chicken", "g"),
    "Pork": ("Pork", "g"),
    "Fish": ("Fish", "g"),
    "Prawns": ("Prawns", "g"),
    "Cuttlefish": ("Cuttlefish", "g"),
    "French Fries": ("French Fries", "g"),
    "Cheese": ("Cheese", "pcs"),
}

warnings: list[str] = []


def warn(msg: str) -> None:
    warnings.append(msg)
    print(f"  ! {msg}")


def num(v):
    """Coerce a cell to a JSON number (int when integral), else None."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v) if v.is_integer() else round(v, 4)
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    try:
        f = float(s)
    except ValueError:
        return None
    return int(f) if f.is_integer() else round(f, 4)


def barcode_str(v):
    """Barcodes are stored as integers in the sheet — read them exactly."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    s = str(v).strip()
    return s or None


def title(s: str) -> str:
    return " ".join(w.capitalize() for w in str(s).split())


# ---------------------------------------------------------------------------
# Wine / spirit sheet
# ---------------------------------------------------------------------------
def load_wine():
    wb = openpyxl.load_workbook(WINE_XLSX, data_only=True)
    ws = wb.active

    # Header row = the one containing a "Barcode" cell (row 2 in this sheet).
    header_row = None
    for r in range(1, min(ws.max_row, 6) + 1):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(r, c).value or "").strip().lower() == "barcode":
                header_row = r
                break
        if header_row:
            break
    if not header_row:
        sys.exit("FATAL: could not find the 'Barcode' header row in the wine sheet")

    col = {}
    pour_col = {}
    header_texts = []
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(header_row, c).value
        if raw is None or str(raw).strip() == "":
            continue
        s = str(raw).strip()
        header_texts.append(f"{c}:{s}")
        low = s.lower()
        if "barcode" in low:
            col["barcode"] = c
        elif low in ("name", "item", "item name", "description"):
            col["name"] = c
        elif low.startswith("category"):
            col["category"] = c
        elif "bottle" in low or "can" in low:
            col["bottle"] = c
        elif low == "mm":
            col["mm"] = c
        elif low in ("qty", "quantity") or "qty" in low:
            col["qty"] = c
        else:
            digits = ""
            for ch in s:
                if ch.isdigit():
                    digits += ch
                else:
                    break
            if digits and int(digits) in POUR_SIZES:
                pour_col[int(digits)] = c

    print(f"  wine header row {header_row}: {' | '.join(header_texts)}")
    print(f"  detected cols: {col}")
    print(f"  pour cols: {dict(sorted(pour_col.items(), reverse=True))}")
    for need in ("name", "category", "barcode", "bottle", "mm", "qty"):
        if need not in col:
            warn(f"wine sheet: no '{need}' column detected")
    for size in POUR_SIZES:
        if size not in pour_col:
            warn(f"wine sheet: no {size} ml pour column detected")

    ingredients = []
    menu_items = []
    current_cat = None
    spirits = discretes = 0

    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, col["name"]).value
        name = str(name).strip() if name is not None else ""
        cat_cell = ws.cell(r, col["category"]).value if "category" in col else None
        if cat_cell is not None and str(cat_cell).strip():
            current_cat = title(cat_cell)  # forward-fill merged category cells
        if not name:
            continue

        group = current_cat or "Bar"
        barcode = barcode_str(ws.cell(r, col["barcode"]).value) if "barcode" in col else None
        pours = []
        for size in POUR_SIZES:
            if size in pour_col:
                p = num(ws.cell(r, pour_col[size]).value)
                if p is not None and p > 0:
                    pours.append((size, p))
        bottle = num(ws.cell(r, col["bottle"]).value) if "bottle" in col else None
        mm = num(ws.cell(r, col["mm"]).value) if "mm" in col else None
        qty = num(ws.cell(r, col["qty"]).value) if "qty" in col else None

        if pours:  # ---- spirit sold by the pour ----
            if bottle is not None:
                warn(f"'{name}' has both pour prices and a Bottle/Can price {bottle}; using pours")
            spirits += 1
            ingredients.append(
                {
                    "name": name,
                    "baseUnit": "ml",
                    "barcode": barcode,
                    "openingStock": mm or 0,
                    "reorderLevel": 0,
                    "costPerUnit": 0,
                }
            )
            for size, price in pours:
                menu_items.append(
                    {
                        "name": f"{name} {size}ml",
                        "category": "bar",
                        "station": "bar",
                        "menuGroup": group,
                        "barcode": None,
                        "prices": [{"channel": ch, "price": price} for ch in BAR_CHANNELS],
                        "recipe": [{"ingredient": name, "quantity": size}],
                    }
                )
        elif bottle is not None:  # ---- discrete whole unit ----
            discretes += 1
            ingredients.append(
                {
                    "name": name,
                    "baseUnit": "pcs",
                    "barcode": None,  # barcode lives on the sellable MenuItem below
                    "openingStock": qty or 0,
                    "reorderLevel": 0,
                    "costPerUnit": 0,
                }
            )
            menu_items.append(
                {
                    "name": name,
                    "category": "bar",
                    "station": "bar",
                    "menuGroup": group,
                    "barcode": barcode,
                    "prices": [{"channel": ch, "price": bottle} for ch in BAR_CHANNELS],
                    "recipe": [{"ingredient": name, "quantity": 1}],
                }
            )
        else:
            warn(f"wine row {r} '{name}' has no pour prices and no Bottle/Can price; skipped")

    print(f"  spirits: {spirits}  discrete: {discretes}")
    return ingredients, menu_items


# ---------------------------------------------------------------------------
# Kitchen sheet
# ---------------------------------------------------------------------------
def load_kitchen():
    wb = openpyxl.load_workbook(KITCHEN_XLSX, data_only=True)
    ws = wb.active
    header_row = 1

    CAT_COL, NAME_COL = 1, 3
    price_col = None
    ing_cols = {}  # col idx -> (clean name, unit)
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(header_row, c).value
        if raw is None or str(raw).strip() == "":
            continue
        s = str(raw).strip()
        if s.lower() == "price" and price_col is None:
            price_col = c
        elif s in KITCHEN_ING:
            ing_cols[c] = KITCHEN_ING[s]

    if price_col is None:
        warn("kitchen sheet: no 'Price' column detected; defaulting to col 5")
        price_col = 5
    print(f"  kitchen price col: {price_col}")
    print(f"  kitchen ingredient cols: { {c: v[0] for c, v in ing_cols.items()} }")
    missing = set(KITCHEN_ING) - {
        str(ws.cell(header_row, c).value or "").strip() for c in ing_cols
    }
    if missing:
        warn(f"kitchen sheet: expected ingredient headers not found: {sorted(missing)}")

    menu_items = []
    used_ing = {}  # clean name -> unit
    ing_qty_range = {}  # clean name -> [min, max]
    current_cat = None
    rows = 0

    for r in range(header_row + 1, ws.max_row + 1):
        cat_cell = ws.cell(r, CAT_COL).value
        if cat_cell is not None and str(cat_cell).strip():
            current_cat = title(cat_cell)
        name = ws.cell(r, NAME_COL).value
        name = str(name).strip() if name is not None else ""
        if not name:
            continue
        price = num(ws.cell(r, price_col).value)
        rows += 1

        recipe = []
        for c, (ing_name, unit) in ing_cols.items():
            q = num(ws.cell(r, c).value)
            if q is not None and q > 0:
                recipe.append({"ingredient": ing_name, "quantity": q})
                used_ing[ing_name] = unit
                lo, hi = ing_qty_range.get(ing_name, (q, q))
                ing_qty_range[ing_name] = (min(lo, q), max(hi, q))

        prices = []
        if price is not None:
            prices = [{"channel": ch, "price": price} for ch in KITCHEN_CHANNELS]
        else:
            warn(f"kitchen '{name}' (row {r}) has no price; item created with no channel prices")

        menu_items.append(
            {
                "name": title(name),
                "category": "food",
                "station": "kitchen",
                "menuGroup": current_cat or "Food",
                "barcode": None,
                "prices": prices,
                "recipe": recipe,
            }
        )

    # Emit every ingredient column that appears in the sheet header, in column
    # order — even one never referenced is a real stock line (opening 0).
    ingredients = []
    for c in sorted(ing_cols):
        ing_name, unit = ing_cols[c]
        ingredients.append(
            {
                "name": ing_name,
                "baseUnit": unit,
                "barcode": None,
                "openingStock": 0,
                "reorderLevel": 0,
                "costPerUnit": 0,
            }
        )
        if ing_name not in used_ing:
            warn(f"kitchen ingredient '{ing_name}' is never used in any recipe")

    print(f"  kitchen dishes: {rows}")
    print("  kitchen ingredient qty ranges (verify pcs vs g):")
    for name in sorted(ing_qty_range):
        lo, hi = ing_qty_range[name]
        unit = used_ing[name]
        print(f"    {name:<14} {unit:<3} min={lo} max={hi}")
    return ingredients, menu_items


def main():
    print("Reading wine/spirit sheet …")
    wine_ings, wine_menu = load_wine()
    print("Reading kitchen sheet …")
    kit_ings, kit_menu = load_kitchen()

    ingredients = wine_ings + kit_ings
    menu_items = wine_menu + kit_menu

    # --- validation --------------------------------------------------------
    seen_ing = {}
    for ing in ingredients:
        if ing["name"] in seen_ing:
            sys.exit(f"FATAL: duplicate ingredient name '{ing['name']}'")
        seen_ing[ing["name"]] = ing

    barcodes = {}
    for ing in ingredients:
        if ing["barcode"]:
            if ing["barcode"] in barcodes:
                sys.exit(f"FATAL: barcode {ing['barcode']} on ingredient '{ing['name']}' collides with {barcodes[ing['barcode']]}")
            barcodes[ing["barcode"]] = f"ingredient:{ing['name']}"
    for mi in menu_items:
        if mi["barcode"]:
            if mi["barcode"] in barcodes:
                sys.exit(f"FATAL: barcode {mi['barcode']} on menu item '{mi['name']}' collides with {barcodes[mi['barcode']]}")
            barcodes[mi["barcode"]] = f"menuItem:{mi['name']}"

    ing_names = set(seen_ing)
    for mi in menu_items:
        for rc in mi["recipe"]:
            if rc["ingredient"] not in ing_names:
                sys.exit(f"FATAL: menu item '{mi['name']}' references unknown ingredient '{rc['ingredient']}'")

    dup_menu = {}
    for mi in menu_items:
        dup_menu[mi["name"]] = dup_menu.get(mi["name"], 0) + 1
    for name, n in dup_menu.items():
        if n > 1:
            warn(f"menu item name '{name}' appears {n} times")

    out = {
        "meta": {
            "generator": "packages/db/scripts/import-xlsx.py",
            "sources": [WINE_XLSX.name, KITCHEN_XLSX.name],
            "note": "Generated from the source workbooks — do not hand-edit; re-run the generator.",
            "counts": {
                "ingredients": len(ingredients),
                "menuItems": len(menu_items),
                "barcodes": len(barcodes),
            },
        },
        "ingredients": ingredients,
        "menuItems": menu_items,
    }

    OUT_JSON.write_text(json.dumps(out, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    print("\n=== summary ===")
    print(f"ingredients : {len(ingredients)}")
    print(f"menu items  : {len(menu_items)}")
    print(f"barcodes    : {len(barcodes)}")
    total_recipes = sum(len(mi["recipe"]) for mi in menu_items)
    total_prices = sum(len(mi["prices"]) for mi in menu_items)
    print(f"recipe rows : {total_recipes}")
    print(f"price rows  : {total_prices}")
    print(f"written     : {OUT_JSON}")
    if warnings:
        print(f"\n{len(warnings)} warning(s) above — review before seeding.")
    # Spot-check: first few barcodes so a human can eyeball exactness.
    print("\nsample barcodes:")
    for code, owner in list(barcodes.items())[:6]:
        print(f"  {code}  ->  {owner}")


if __name__ == "__main__":
    main()
